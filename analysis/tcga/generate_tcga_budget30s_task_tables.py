#!/usr/bin/env python
"""Generate TCGA pca100 budget30s ICL-vs-classical metric tables.

The hparam raw CSV/XLSX files are used only to identify the selected
prediction payload for each method/task/search run. AUROC and macro F1 are
recomputed from ``best_candidate_test_prediction.pkl``.
"""

from __future__ import annotations

import argparse
import io
import os
import pickle
import sys
from contextlib import redirect_stdout
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from scipy.stats import ttest_rel
from sklearn.metrics import f1_score, roc_auc_score


REPO_ROOT = Path(__file__).resolve().parents[2]
FMICL_ROOT = REPO_ROOT / "fmicl"
for path in (REPO_ROOT, FMICL_ROOT):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from tcga_ancestral_continuum_utils import load_expression_data  # noqa: E402


OUTPUT_DIR = REPO_ROOT / "results" / "tables" / "budget30s_task_tables"
METHOD_DIR = Path(
    os.environ.get(
        "FM_ICL_METHOD_RESULT_DIR",
        "/nfs/home/wli66/tabular_transformer_transfer_learning/data_analytics/tcga/budget_curve_method_excels",
    )
)
LABEL_CACHE_DIR = (
    Path(
        os.environ.get(
            "FM_ICL_LABEL_CACHE_ROOT",
            "/lustre/isaac24/scratch/wli66/tabular_transformer_transfer_learning/Data/TCGA/ancestral_continuum",
        )
    )
)

BUDGET = "budget30s"
YEAR = 2
CANCERS = ["KIRC", "UCEC", "BRCA"]
TARGETS = ["OS", "PFI"]
OMICS = ["mRNA", "MicroRNA", "Methylation"]

TASK_BASE_ORDER = [
    ("KIRC", "OS"),
    ("KIRC", "PFI"),
    ("UCEC", "OS"),
    ("UCEC", "PFI"),
    ("BRCA", "OS"),
    ("BRCA", "PFI"),
]

METHOD_DISPLAY = {
    "icl_min": "FM-ICL (AFR)",
    "icl_mixed": "FM-ICL (Mix)",
    "elasticnet_min": "Elastic Net (AFR)",
    "elasticnet_mixed": "Elastic Net (Mix)",
    "randomforest_min": "Random Forest (AFR)",
    "randomforest_mixed": "Random Forest (Mix)",
    "xgboost_min": "XGBoost (AFR)",
    "xgboost_mixed": "XGBoost (Mix)",
}
METHOD_ORDER = list(METHOD_DISPLAY)

CLASSICAL_METHODS = [
    ("Elastic Net", "elasticnet_min", "elasticnet_mixed"),
    ("Random Forest", "randomforest_min", "randomforest_mixed"),
    ("XGBoost", "xgboost_min", "xgboost_mixed"),
]

LABEL_LOOKUPS: dict[tuple[str, str], dict[str, int]] = {}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate TCGA pca100 budget30s ICL-vs-classical metric tables."
    )
    parser.add_argument("--method-dir", type=Path, default=METHOD_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--budget", default=BUDGET)
    parser.add_argument("--year", type=int, default=YEAR)
    parser.add_argument("--cancers", nargs="+", default=CANCERS, choices=CANCERS)
    parser.add_argument("--targets", nargs="+", default=TARGETS, choices=TARGETS)
    parser.add_argument("--omics", nargs="+", default=OMICS, choices=OMICS)
    parser.add_argument("--methods", nargs="+", default=METHOD_ORDER, choices=METHOD_ORDER)
    parser.add_argument(
        "--alpha",
        type=float,
        default=0.05,
        help="Significance threshold used for the check mark in t-test cells.",
    )
    return parser.parse_args()


def task_name(cancer: str, target: str, omics: str) -> str:
    return f"{cancer}-{target}-{omics}"


def task_order(cancers: Iterable[str], targets: Iterable[str], omics_list: Iterable[str]) -> list[str]:
    requested = {
        (cancer, target, omics)
        for cancer in cancers
        for target in targets
        for omics in omics_list
    }
    ordered: list[str] = []
    for omics in OMICS:
        for cancer, target in TASK_BASE_ORDER:
            if (cancer, target, omics) in requested:
                ordered.append(task_name(cancer, target, omics))
    return ordered


def method_input_path(method_dir: Path, method: str) -> Path:
    csv_path = method_dir / f"budget_curves_raw_pca100_{method}.csv"
    if csv_path.is_file():
        return csv_path
    xlsx_path = method_dir / f"budget_curves_raw_pca100_{method}.xlsx"
    if xlsx_path.is_file():
        return xlsx_path
    raise FileNotFoundError(f"Missing hparam file for {method}: {csv_path} or {xlsx_path}")


def read_method_rows(method_dir: Path, method: str) -> pd.DataFrame:
    path = method_input_path(method_dir, method)
    print(f"[read] {method}: {path}")
    if path.suffix.lower() in {".xlsx", ".xls"}:
        data = pd.read_excel(path)
    else:
        data = pd.read_csv(path)

    required = {"method", "cancer", "omics", "target", "budget", "search_run", "result_dir"}
    missing = sorted(required.difference(data.columns))
    if missing:
        raise ValueError(f"{path} is missing required columns: {missing}")
    return data


def build_label_lookup(omics: str, target: str, year: int) -> dict[str, int]:
    key = (omics, target)
    if key in LABEL_LOOKUPS:
        return LABEL_LOOKUPS[key]

    with redirect_stdout(io.StringIO()):
        expression_data = load_expression_data(
            omics,
            str(LABEL_CACHE_DIR),
            year=year,
            target=target,
        )
    samples = [str(sample_id) for sample_id in expression_data["Samples"]]
    labels = [int(label) for label in expression_data["Y"]]
    LABEL_LOOKUPS[key] = dict(zip(samples, labels))
    return LABEL_LOOKUPS[key]


def normalize_predictions(pred_block: object) -> np.ndarray:
    pred_arr = np.asarray(pred_block, dtype=float)
    if pred_arr.ndim == 1:
        pred_arr = np.column_stack((1.0 - pred_arr, pred_arr))
    if pred_arr.ndim != 2 or pred_arr.shape[1] < 2:
        raise ValueError(f"invalid prediction shape: {pred_arr.shape}")
    return pred_arr


def metrics_from_prediction_payload(payload_path: Path, label_lookup: dict[str, int]) -> dict[str, float] | None:
    try:
        with payload_path.open("rb") as handle:
            payload = pickle.load(handle)
    except Exception:
        return None

    fold_aurocs: list[float] = []
    fold_f1s: list[float] = []
    for pred_block, sample_block in zip(payload.get("pred", []), payload.get("sample_name", [])):
        try:
            pred_arr = normalize_predictions(pred_block)
        except Exception:
            continue

        sample_ids = [str(sample_id) for sample_id in sample_block]
        limit = min(len(sample_ids), pred_arr.shape[0])
        if limit == 0:
            continue
        sample_ids = sample_ids[:limit]
        pred_arr = pred_arr[:limit]

        keep_idx = [idx for idx, sample_id in enumerate(sample_ids) if sample_id in label_lookup]
        if not keep_idx:
            continue

        y_true = np.asarray([label_lookup[sample_ids[idx]] for idx in keep_idx], dtype=int)
        if len(np.unique(y_true)) < 2:
            continue
        try:
            fold_auroc = float(roc_auc_score(y_true, pred_arr[keep_idx, -1]))
            y_pred = pred_arr[keep_idx].argmax(axis=1)
            fold_f1 = float(f1_score(y_true, y_pred, average="macro"))
        except Exception:
            continue
        if not np.isnan(fold_auroc) and not np.isnan(fold_f1):
            fold_aurocs.append(fold_auroc)
            fold_f1s.append(fold_f1)

    if not fold_aurocs or not fold_f1s:
        return None
    return {
        "auroc": float(np.nanmean(fold_aurocs)),
        "f1": float(np.nanmean(fold_f1s)),
    }


def collect_run_level_metrics(args: argparse.Namespace) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    label_cache: dict[tuple[str, str], dict[str, int]] = {}

    for method in args.methods:
        data = read_method_rows(args.method_dir, method)
        data = data.loc[
            data["budget"].astype(str).eq(args.budget)
            & data["cancer"].astype(str).isin(args.cancers)
            & data["target"].astype(str).isin(args.targets)
            & data["omics"].astype(str).isin(args.omics)
        ].copy()
        data["search_run"] = pd.to_numeric(data["search_run"], errors="coerce").astype("Int64")

        total = data.shape[0]
        for idx, record in enumerate(data.to_dict("records"), start=1):
            if idx % 100 == 0:
                print(f"  [progress] {method}: {idx}/{total}")
            cancer = str(record["cancer"])
            target = str(record["target"])
            omics = str(record["omics"])
            key = (omics, target)
            if key not in label_cache:
                label_cache[key] = build_label_lookup(omics, target, args.year)

            prediction_path = Path(str(record["result_dir"])) / "best_candidate_test_prediction.pkl"
            metrics = (
                metrics_from_prediction_payload(prediction_path, label_cache[key])
                if prediction_path.is_file()
                else None
            )
            rows.append(
                {
                    "method": method,
                    "method_label": METHOD_DISPLAY[method],
                    "group": "AFR" if method.endswith("_min") else "Mix",
                    "cancer": cancer,
                    "target": target,
                    "omics": omics,
                    "task": task_name(cancer, target, omics),
                    "budget": str(record["budget"]),
                    "search_run": int(record["search_run"]) if not pd.isna(record["search_run"]) else np.nan,
                    "auroc": metrics["auroc"] if metrics is not None else None,
                    "f1": metrics["f1"] if metrics is not None else None,
                    "prediction_path": str(prediction_path),
                    "prediction_found": prediction_path.is_file(),
                    "source_result_dir": str(record["result_dir"]),
                }
            )

    return pd.DataFrame(rows)


def format_mean_cell(mean_value: float, std_value: float, n_value: int) -> str:
    if n_value <= 0 or np.isnan(mean_value):
        return "NA"
    if np.isnan(std_value):
        std_value = 0.0
    return f"{mean_value:.3f} ± {std_value:.3f}"


def build_mean_table(run_df: pd.DataFrame, tasks: list[str], metric: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary = (
        run_df.dropna(subset=[metric])
        .groupby(["method", "method_label", "task"], observed=True)[metric]
        .agg(mean="mean", std=lambda values: float(values.std(ddof=1)), n="count")
        .reset_index()
    )

    numeric = pd.DataFrame(index=[METHOD_DISPLAY[m] for m in METHOD_ORDER if m in set(run_df["method"])], columns=tasks)
    table = numeric.copy()
    for record in summary.to_dict("records"):
        label = str(record["method_label"])
        task = str(record["task"])
        if task not in table.columns or label not in table.index:
            continue
        mean_value = float(record["mean"])
        std_value = float(record["std"]) if not pd.isna(record["std"]) else np.nan
        n_value = int(record["n"])
        numeric.loc[label, task] = mean_value
        table.loc[label, task] = format_mean_cell(mean_value, std_value, n_value)

    table.index.name = "Method"
    numeric.index.name = "Method"
    return table, numeric.astype(float)


def pvalue_text(p_value: float | None) -> str:
    if p_value is None or np.isnan(p_value):
        return "NA"
    if p_value < 1e-4:
        return f"{p_value:.2e}"
    return f"{p_value:.3f}"


def one_paired_ttest(
    run_df: pd.DataFrame,
    task: str,
    icl_method: str,
    baseline_method: str,
    metric: str,
    alpha: float,
) -> tuple[str, dict[str, object]]:
    subset = run_df.loc[
        run_df["task"].eq(task)
        & run_df["method"].isin([icl_method, baseline_method])
        & run_df[metric].notna(),
        ["method", "search_run", metric],
    ].copy()
    wide = subset.pivot_table(index="search_run", columns="method", values=metric, aggfunc="mean")
    if icl_method not in wide.columns or baseline_method not in wide.columns:
        n_pairs = 0
        icl_mean = baseline_mean = mean_diff = t_stat = p_value = np.nan
    else:
        paired = wide[[icl_method, baseline_method]].dropna()
        n_pairs = int(paired.shape[0])
        icl_values = paired[icl_method].astype(float)
        baseline_values = paired[baseline_method].astype(float)
        icl_mean = float(icl_values.mean()) if n_pairs else np.nan
        baseline_mean = float(baseline_values.mean()) if n_pairs else np.nan
        mean_diff = icl_mean - baseline_mean if n_pairs else np.nan
        if n_pairs >= 2 and not np.allclose(icl_values - baseline_values, 0.0):
            test_result = ttest_rel(icl_values, baseline_values, nan_policy="omit")
            t_stat = float(test_result.statistic)
            p_value = float(test_result.pvalue)
        elif n_pairs >= 2:
            t_stat = 0.0
            p_value = 1.0
        else:
            t_stat = np.nan
            p_value = np.nan

    significant = bool(
        n_pairs >= 2
        and not np.isnan(mean_diff)
        and not np.isnan(p_value)
        and p_value < alpha
    )
    icl_better = bool(significant and mean_diff > 0)
    baseline_better = bool(significant and mean_diff < 0)
    if icl_better:
        marker = "✓"
    elif baseline_better:
        marker = "x"
    else:
        marker = "-"
    cell = f"{marker} (p = {pvalue_text(p_value)})"
    details = {
        "task": task,
        "metric": metric,
        "icl_method": icl_method,
        "baseline_method": baseline_method,
        "n_pairs": n_pairs,
        "icl_mean": icl_mean,
        "baseline_mean": baseline_mean,
        "mean_diff": mean_diff,
        "t_stat": t_stat,
        "p_value": p_value,
        "marker": marker,
        "icl_better": icl_better,
        "baseline_better": baseline_better,
    }
    return cell, details


def build_ttest_table(
    run_df: pd.DataFrame,
    tasks: list[str],
    metric: str,
    alpha: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    table = pd.DataFrame(index=[f"ICL vs {name}" for name, _, _ in CLASSICAL_METHODS], columns=tasks)
    details: list[dict[str, object]] = []

    for classical_name, min_method, mixed_method in CLASSICAL_METHODS:
        row_label = f"ICL vs {classical_name}"
        for task in tasks:
            afr_cell, afr_details = one_paired_ttest(run_df, task, "icl_min", min_method, metric, alpha)
            mix_cell, mix_details = one_paired_ttest(run_df, task, "icl_mixed", mixed_method, metric, alpha)

            table.loc[row_label, task] = f"AFR: {afr_cell}\nMix: {mix_cell}"
            afr_details.update({"comparison": row_label, "group": "AFR"})
            mix_details.update({"comparison": row_label, "group": "Mix"})
            details.extend([afr_details, mix_details])

    table.index.name = "Comparison"
    return table, pd.DataFrame(details)


def write_excel_tables(
    mean_table: pd.DataFrame,
    mean_numeric: pd.DataFrame,
    ttest_table: pd.DataFrame,
    output_dir: Path,
    metric: str,
) -> None:
    mean_path = output_dir / f"mean_{metric}_comparison_table.xlsx"
    ttest_path = output_dir / f"paired_ttest_{metric}_table.xlsx"
    mean_table.to_excel(mean_path)
    ttest_table.to_excel(ttest_path)

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    workbook = load_workbook(mean_path)
    worksheet = workbook.active
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, task in enumerate(mean_table.columns, start=2):
        values = mean_numeric[task].dropna()
        if values.empty:
            continue
        best = float(values.max())
        for row_idx, method_label in enumerate(mean_table.index, start=2):
            value = mean_numeric.loc[method_label, task]
            if not pd.isna(value) and np.isclose(float(value), best):
                worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    worksheet.column_dimensions["A"].width = 26
    for col_cells in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
        worksheet.column_dimensions[col_cells[0].column_letter].width = 20
    workbook.save(mean_path)

    workbook = load_workbook(ttest_path)
    worksheet = workbook.active
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.column_dimensions["A"].width = 24
    for col_cells in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
        worksheet.column_dimensions[col_cells[0].column_letter].width = 22
    workbook.save(ttest_path)


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    tasks = task_order(args.cancers, args.targets, args.omics)
    print(f"[info] task_count={len(tasks)}")
    print("[info] recomputing AUROC and macro F1 from prediction payloads")

    run_df = collect_run_level_metrics(args)
    run_path = args.output_dir / "run_level_recomputed_metrics.csv"
    run_df.to_csv(run_path, index=False)
    print(f"[saved] {run_path}")

    for metric in ("auroc", "f1"):
        mean_table, mean_numeric = build_mean_table(run_df, tasks, metric)
        mean_csv = args.output_dir / f"mean_{metric}_comparison_table.csv"
        mean_table.to_csv(mean_csv)
        print(f"[saved] {mean_csv}")

        ttest_table, ttest_details = build_ttest_table(run_df, tasks, metric, args.alpha)
        ttest_csv = args.output_dir / f"paired_ttest_{metric}_table.csv"
        ttest_table.to_csv(ttest_csv)
        print(f"[saved] {ttest_csv}")

        details_path = args.output_dir / f"paired_ttest_{metric}_details.csv"
        ttest_details.to_csv(details_path, index=False)
        print(f"[saved] {details_path}")

        write_excel_tables(mean_table, mean_numeric, ttest_table, args.output_dir, metric)
        print(f"[saved] {args.output_dir / f'mean_{metric}_comparison_table.xlsx'}")
        print(f"[saved] {args.output_dir / f'paired_ttest_{metric}_table.xlsx'}")

        print(f"[check] {metric}_mean_shape={mean_table.shape}")
        print(f"[check] {metric}_ttest_shape={ttest_table.shape}")
        print(f"[check] valid_{metric}={int(run_df[metric].notna().sum())}/{run_df.shape[0]}")


if __name__ == "__main__":
    main()
